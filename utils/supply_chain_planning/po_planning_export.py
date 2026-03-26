# utils/supply_chain_planning/po_planning_export.py

"""
Export Module for PO Planning — multi-sheet Excel export with formatting.

v1.1: Added openpyxl formatting:
- Column width auto-fit
- Number/currency formatting
- Conditional formatting (urgency row colors)
- Freeze panes (header row fixed when scrolling)
- Auto-filter on all sheets
- Header row styling (bold, blue background)
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
import logging

from .planning_constants import URGENCY_LEVELS

logger = logging.getLogger(__name__)

HEADER_FILL_HEX = '2E75B6'
HEADER_FONT_COLOR = 'FFFFFF'
URGENCY_COLORS = {
    'OVERDUE': 'F2DCDB', 'CRITICAL': 'F2DCDB',
    'URGENT': 'FDE9D9', 'THIS_WEEK': 'FFFFCC', 'PLANNED': 'D6EAFF',
}
DEFAULT_COL_WIDTHS = {
    'Urgency': 14, 'Source': 12, 'Vendor': 25, 'Vendor Code': 12,
    'Code': 14, 'Product': 30, 'Brand': 12, 'Pkg Size': 10,
    'UOM': 8, 'GAP Shortage': 14, 'Pending PO': 12, 'Net Need': 12,
    'Order Qty': 12, 'MOQ': 10, 'SPQ': 10,
    'Unit Price': 12, 'Unit Price (USD)': 14, 'Currency': 10,
    'Line Value (USD)': 16, 'VAT %': 8,
    'Price Source': 12, 'Costbook #': 14, 'Last PO #': 14,
    'Lead Time (days)': 14, 'LT Source': 12, 'Reliability': 12,
    'Demand Date': 14, 'Must Order By': 14, 'Expected Arrival': 14,
    'Days to Order': 12, 'Overdue?': 10,
    'Trade Term': 12, 'Payment Term': 14, 'Shipping': 12,
    'Vendor Match Notes': 30, 'Qty Notes': 25,
    'Location': 14, 'PO Lines': 10, 'Total Value (USD)': 16,
    'Max Urgency': 14, 'Priority': 10,
    'Metric': 35, 'Value': 40, 'Reason': 40, 'Shortage Qty': 14,
}


def _apply_sheet_formatting(ws, df, col_widths=None, currency_cols=None,
                            number_cols=None, date_cols=None, urgency_col=None):
    """Apply professional formatting to a worksheet."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    thin = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
    )
    hdr_font = Font(name='Arial', bold=True, color=HEADER_FONT_COLOR, size=10)
    hdr_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dfont = Font(name='Arial', size=10)

    nrows, ncols = len(df), len(df.columns)

    for ci in range(1, ncols + 1):
        c = ws.cell(row=1, column=ci)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, thin

    ws.freeze_panes = 'A2'
    if nrows > 0:
        ws.auto_filter.ref = ws.dimensions

    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}
    widths = dict(DEFAULT_COL_WIDTHS)
    if col_widths:
        widths.update(col_widths)
    for name, ci in col_map.items():
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, max(len(str(name)) + 2, 12))

    curr_set = set(currency_cols or [])
    num_set = set(number_cols or [])
    date_set = set(date_cols or [])

    for ri in range(2, nrows + 2):
        for name, ci in col_map.items():
            c = ws.cell(row=ri, column=ci)
            c.font, c.border = dfont, thin
            c.alignment = Alignment(vertical='center')
            if name in curr_set:
                c.number_format = '$#,##0.00'
            elif name in num_set:
                c.number_format = '#,##0'
            elif name in date_set:
                c.number_format = 'YYYY-MM-DD'

    if urgency_col and urgency_col in col_map:
        uci = col_map[urgency_col]
        for ri in range(2, nrows + 2):
            val = ws.cell(row=ri, column=uci).value
            if not val:
                continue
            vu = str(val).upper()
            for uk, hx in URGENCY_COLORS.items():
                if uk in vu:
                    fill = PatternFill(start_color=hx, end_color=hx, fill_type='solid')
                    for ci in range(1, ncols + 1):
                        ws.cell(row=ri, column=ci).fill = fill
                    break


# =============================================================================
# MAIN EXPORT
# =============================================================================

def _get_scope_display(result):
    """Get human-readable scope label for export."""
    inp = getattr(result, 'input_summary', None) or {}
    scope = inp.get('filter_scope', 'full')
    scope_info = inp.get('scope_info', {})
    if scope == 'filtered' and scope_info.get('has_filter'):
        return f"Filtered ({scope_info.get('scope_label', 'custom')})"
    return 'Full (all products)'


def export_po_suggestions_to_excel(result, gap_summary=None):
    """Export PO suggestions to formatted Excel workbook."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        _write_summary_sheet(writer, result, gap_summary)
        _write_po_lines_sheet(writer, result)
        _write_vendor_sheet(writer, result)
        if result.has_unmatched():
            _write_unmatched_sheet(writer, result)
        if hasattr(result, 'skipped_items') and result.skipped_items:
            _write_skipped_sheet(writer, result)
    buffer.seek(0)
    return buffer


def _write_summary_sheet(writer, result, gap_summary):
    metrics = result.get_summary()
    data = [
        ['PO Planning Summary', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Strategy', metrics.get('strategy', 'CHEAPEST')],
        ['Scope', _get_scope_display(result)],
        ['', ''],
        ['--- PO SUGGESTIONS ---', ''],
        ['Total PO Lines', metrics.get('total_po_lines', 0)],
        ['FG Trading Lines', metrics.get('fg_lines', 0)],
        ['Raw Material Lines', metrics.get('raw_lines', 0)],
        ['Total Vendors', metrics.get('total_vendors', 0)],
        ['Total Value (USD)', f"${metrics.get('total_value_usd', 0):,.2f}"],
        ['', ''],
        ['--- URGENCY ---', ''],
    ]
    dist = metrics.get('urgency_distribution', {})
    for lk in ['OVERDUE', 'CRITICAL', 'URGENT', 'THIS_WEEK', 'PLANNED']:
        cnt = dist.get(lk, 0)
        if cnt > 0:
            cfg = URGENCY_LEVELS.get(lk, {})
            data.append([f"  {cfg.get('label', lk)}", cnt])

    data += [['', ''], ['--- VENDOR MATCHING ---', ''],
             ['Matched to Vendor', metrics.get('total_po_lines', 0)],
             ['Skipped (Pending PO Covers)', metrics.get('skipped_count', 0)],
             ['No Vendor Found', metrics.get('unmatched_count', 0)]]

    # Data Reconciliation
    recon = metrics.get('reconciliation', {})
    if recon:
        data += [['', ''], ['--- DATA RECONCILIATION ---', '']]
        scope_label = recon.get('filter_scope', 'full')
        if scope_label == 'filtered':
            data.append(['Scope', f"Filtered — see scope info above"])
        data.append(['Input from GAP', recon.get('total_input', 0)])
        data.append([f'  FG Trading', recon.get('input_fg', 0)])
        data.append([f'  Raw Material', recon.get('input_raw', 0)])
        data.append(['', ''])
        data.append(['→ PO Lines Created', recon.get('matched', 0)])
        data.append(['→ Skipped (Pending PO)', recon.get('skipped_pending_po', 0)])
        data.append(['→ No Vendor Found', recon.get('unmatched', 0)])
        data.append(['→ Validation Skipped', recon.get('input_skipped_validation', 0)])
        scope_removed = recon.get('scope_removed', 0)
        if scope_removed > 0:
            data.append([f'→ Scope Filtered Out', scope_removed])
            data.append([f'    FG removed by scope', recon.get('scope_removed_fg', 0)])
            data.append([f'    Raw removed by scope', recon.get('scope_removed_raw', 0)])
        data.append(['→ Processing Errors', recon.get('processing_errors', 0)])
        data.append(['', ''])
        data.append(['Total Accounted', recon.get('total_accounted', 0)])
        balanced = '✅ YES' if recon.get('is_balanced', False) else f"❌ NO (diff={recon.get('discrepancy', '?')})"
        data.append(['Balanced?', balanced])

    # GAP Filter Warnings
    inp = getattr(result, 'input_summary', None) or {}
    filter_review = inp.get('filter_review', {})
    filter_items = filter_review.get('items', [])
    if filter_items:
        data += [['', ''], ['--- GAP FILTER WARNINGS ---', '']]
        data.append(['Filter Status', filter_review.get('summary_text', '')])
        for item in filter_items:
            risk_icon = {'HIGH': '🔴', 'MEDIUM': '🟡', 'INFO': 'ℹ️'}.get(item.get('risk'), '')
            data.append([
                f"  {risk_icon} {item.get('label', '')} ({item.get('risk', '')})",
                item.get('consequence', '')
            ])
        data.append(['', ''])
        data.append(['⚠️ NOTE', 'PO suggestions based on incomplete GAP filters — verify before ordering'])

    errors = metrics.get('processing_errors', [])
    if errors:
        data += [['', ''], ['--- PROCESSING ERRORS ---', '']]
        for err in errors[:10]:
            data.append(['Error', str(err)[:100]])

    if gap_summary:
        data += [['', ''], ['--- SOURCE: SCM GAP ---', '']]
        for key in ['fg_shortage_items', 'po_fg_count', 'po_raw_count', 'at_risk_value']:
            if key in gap_summary:
                data.append([key.replace('_', ' ').title(), gap_summary[key]])

    df = pd.DataFrame(data, columns=['Metric', 'Value'])
    df.to_excel(writer, sheet_name='Summary', index=False)
    _apply_sheet_formatting(writer.sheets['Summary'], df)


def _write_po_lines_sheet(writer, result):
    lines_df = result.get_all_lines_df()
    if lines_df.empty:
        pd.DataFrame({'Note': ['No PO suggestions']}).to_excel(writer, sheet_name='PO Lines', index=False)
        return

    lines_df = lines_df.sort_values(['urgency_priority', 'vendor_name', 'pt_code']).reset_index(drop=True)

    columns = [
        'urgency_level', 'shortage_source', 'vendor_name', 'vendor_code',
        'pt_code', 'product_name', 'brand', 'package_size', 'standard_uom',
        'shortage_qty', 'pending_po_qty', 'net_shortage_qty',
        'suggested_qty', 'moq', 'spq', 'moq_applied', 'spq_applied', 'excess_qty',
        'unit_price', 'unit_price_usd', 'currency_code', 'line_value_usd', 'vat_percent',
        'price_source', 'costbook_number', 'last_po_number',
        'lead_time_days', 'lead_time_source', 'lead_time_notes', 'vendor_reliability',
        'demand_date', 'must_order_by', 'expected_arrival',
        'days_until_must_order', 'is_overdue',
        'trade_term', 'payment_term', 'shipping_mode',
        'match_notes', 'quantity_notes',
    ]
    available = [c for c in columns if c in lines_df.columns]
    export_df = lines_df[available].copy()

    rename = {
        'urgency_level': 'Urgency', 'shortage_source': 'Source',
        'vendor_name': 'Vendor', 'vendor_code': 'Vendor Code',
        'pt_code': 'Code', 'product_name': 'Product',
        'brand': 'Brand', 'package_size': 'Pkg Size', 'standard_uom': 'UOM',
        'shortage_qty': 'GAP Shortage', 'pending_po_qty': 'Pending PO',
        'net_shortage_qty': 'Net Need', 'suggested_qty': 'Order Qty',
        'moq': 'MOQ', 'spq': 'SPQ', 'moq_applied': 'MOQ Applied',
        'spq_applied': 'SPQ Applied', 'excess_qty': 'Excess Qty',
        'unit_price': 'Unit Price', 'unit_price_usd': 'Unit Price (USD)',
        'currency_code': 'Currency', 'line_value_usd': 'Line Value (USD)',
        'vat_percent': 'VAT %', 'price_source': 'Price Source',
        'costbook_number': 'Costbook #', 'last_po_number': 'Last PO #',
        'lead_time_days': 'Lead Time (days)', 'lead_time_source': 'LT Source',
        'lead_time_notes': 'LT Breakdown',
        'vendor_reliability': 'Reliability',
        'demand_date': 'Demand Date', 'must_order_by': 'Must Order By',
        'expected_arrival': 'Expected Arrival', 'days_until_must_order': 'Days to Order',
        'is_overdue': 'Overdue?', 'trade_term': 'Trade Term',
        'payment_term': 'Payment Term', 'shipping_mode': 'Shipping',
        'match_notes': 'Vendor Match Notes', 'quantity_notes': 'Qty Notes',
    }
    export_df.rename(columns=rename, inplace=True)

    for col in ['MOQ Applied', 'SPQ Applied', 'Overdue?']:
        if col in export_df.columns:
            export_df[col] = export_df[col].apply(lambda x: 'Yes' if x else 'No')

    export_df.to_excel(writer, sheet_name='PO Lines', index=False)
    _apply_sheet_formatting(
        writer.sheets['PO Lines'], export_df,
        currency_cols=['Unit Price', 'Unit Price (USD)', 'Line Value (USD)'],
        number_cols=['GAP Shortage', 'Pending PO', 'Net Need', 'Order Qty',
                     'MOQ', 'SPQ', 'Excess Qty', 'Lead Time (days)', 'Days to Order'],
        date_cols=['Demand Date', 'Must Order By', 'Expected Arrival'],
        urgency_col='Urgency',
    )


def _write_vendor_sheet(writer, result):
    vendor_df = result.get_vendor_summary_df()
    if vendor_df.empty:
        pd.DataFrame({'Note': ['No vendor data']}).to_excel(writer, sheet_name='By Vendor', index=False)
        return
    vendor_df = vendor_df.sort_values('max_urgency_priority').reset_index(drop=True)
    rename = {
        'vendor_name': 'Vendor', 'vendor_code': 'Code',
        'vendor_location_type': 'Location', 'vendor_reliability': 'Reliability',
        'total_lines': 'PO Lines', 'total_value_usd': 'Total Value (USD)',
        'primary_currency': 'Currency', 'max_urgency_level': 'Max Urgency',
        'max_urgency_priority': 'Priority',
        'trade_term': 'Trade Term', 'payment_term': 'Payment Term',
    }
    export_df = vendor_df.rename(columns=rename)
    export_df.to_excel(writer, sheet_name='By Vendor', index=False)
    _apply_sheet_formatting(
        writer.sheets['By Vendor'], export_df,
        currency_cols=['Total Value (USD)'], number_cols=['PO Lines', 'Priority'],
        urgency_col='Max Urgency',
    )


def _write_unmatched_sheet(writer, result):
    unmatched_df = result.get_unmatched_df()
    if unmatched_df.empty:
        return
    rename = {
        'pt_code': 'Code', 'product_name': 'Product', 'brand': 'Brand',
        'shortage_source': 'Source', 'shortage_qty': 'Shortage Qty',
        'uom': 'UOM', 'reason': 'Reason',
    }
    export_df = unmatched_df.rename(columns=rename)
    export_df.to_excel(writer, sheet_name='Unmatched', index=False)
    _apply_sheet_formatting(writer.sheets['Unmatched'], export_df, number_cols=['Shortage Qty'])


def _write_skipped_sheet(writer, result):
    """Write skipped items sheet — products where pending PO already covers shortage."""
    skipped_df = result.get_skipped_df()
    if skipped_df.empty:
        return
    rename = {
        'pt_code': 'Code', 'product_name': 'Product', 'brand': 'Brand',
        'shortage_source': 'Source', 'shortage_qty': 'Shortage Qty',
        'pending_po_qty': 'Pending PO Qty', 'net_shortage_qty': 'Net Shortage',
        'uom': 'UOM', 'vendor_name': 'Vendor', 'reason': 'Reason',
    }
    export_df = skipped_df.rename(columns=rename)
    export_df.to_excel(writer, sheet_name='Skipped (PO Covers)', index=False)
    _apply_sheet_formatting(
        writer.sheets['Skipped (PO Covers)'], export_df,
        number_cols=['Shortage Qty', 'Pending PO Qty', 'Net Shortage'],
    )


def get_po_export_filename(prefix="po_suggestions"):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{prefix}_{ts}.xlsx"