# utils/inventory_quality/common.py
"""
Common utilities for Inventory Quality module
Formatting, validation, UI helpers, and constants

Version: 1.0.0
"""

import logging
from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, Union, Any
from io import BytesIO

import pandas as pd
import streamlit as st

# Timezone support
try:
    from zoneinfo import ZoneInfo
    VN_TIMEZONE = ZoneInfo('Asia/Ho_Chi_Minh')
except ImportError:
    try:
        import pytz
        VN_TIMEZONE = pytz.timezone('Asia/Ho_Chi_Minh')
    except ImportError:
        VN_TIMEZONE = None
        logging.warning("No timezone library available. Using system timezone.")

logger = logging.getLogger(__name__)


# ==================== Constants ====================

class InventoryQualityConstants:
    """Inventory Quality specific constants"""
    DEFAULT_PAGE_SIZE = 50
    MAX_PAGE_SIZE = 500
    QUANTITY_DECIMALS = 4
    CURRENCY_DECIMALS = 2
    
    # Categories
    CATEGORY_GOOD = 'GOOD'
    CATEGORY_QUARANTINE = 'QUARANTINE'
    CATEGORY_DEFECTIVE = 'DEFECTIVE'
    
    # Category display
    CATEGORY_DISPLAY = {
        'GOOD': '📗 Good',
        'QUARANTINE': '📙 Quarantine',
        'DEFECTIVE': '📕 Defective'
    }
    
    CATEGORY_COLORS = {
        'GOOD': '#28a745',
        'QUARANTINE': '#ffc107', 
        'DEFECTIVE': '#dc3545'
    }
    
    # Defect types - includes new partial QC defect types
    DEFECT_TYPES = {
        # Legacy defect types
        'QC_PENDING': 'QC Pending',
        'QC_FAILED': 'QC Failed',
        'DAMAGED': 'Damaged',
        'EXPIRED': 'Expired',
        'CONTAMINATED': 'Contaminated',
        # New partial QC defect types
        'VISUAL': '🔍 Visual Defect',
        'DIMENSIONAL': '📏 Dimensional',
        'FUNCTIONAL': '⚙️ Functional',
        'CONTAMINATION': '🧪 Contamination',
        'PACKAGING': '📦 Packaging',
        'OTHER': '❓ Other'
    }
    
    # Source types - includes partial QC source
    SOURCE_TYPES = {
        'Opening Balance': 'Opening Balance',
        'Production': 'Production',
        'Production Return': 'Production Return',
        'Production (Pending QC)': 'Production (Pending QC)',
        'Production (Failed QC)': 'Production (Failed QC)',
        'Production (Partial QC Failed)': 'Production (Partial QC Failed)',  # NEW
        'Material Return (Damaged)': 'Material Return (Damaged)',
        'Direct Stock-In': 'Direct Stock-In'
    }
    
    # Period report presets
    PERIOD_PRESETS = {
        'this_month': 'This Month',
        'last_month': 'Last Month',
        'this_quarter': 'This Quarter',
        'last_quarter': 'Last Quarter',
        'this_year': 'This Year',
        'last_year': 'Last Year',
        'custom': 'Custom',
    }


# ==================== Timezone Helpers ====================

def get_vietnam_now() -> datetime:
    """Get current datetime in Vietnam timezone (UTC+7)"""
    if VN_TIMEZONE:
        return datetime.now(VN_TIMEZONE)
    return datetime.now()


def get_vietnam_today() -> date:
    """Get current date in Vietnam timezone (UTC+7)"""
    if VN_TIMEZONE:
        return datetime.now(VN_TIMEZONE).date()
    return date.today()


# ==================== Formatting Helpers ====================

def format_quantity(value: Any, decimals: int = 4) -> str:
    """Format quantity with specified decimal places"""
    if value is None:
        return "-"
    try:
        num = float(value)
        if num == int(num):
            return f"{int(num):,}"
        return f"{num:,.{decimals}f}".rstrip('0').rstrip('.')
    except (ValueError, TypeError):
        return str(value)


def format_currency(value: Any, currency: str = "USD", decimals: int = 2) -> str:
    """Format currency value"""
    if value is None:
        return "-"
    try:
        num = float(value)
        if currency == "VND":
            return f"{int(num):,} ₫"
        return f"${num:,.{decimals}f}"
    except (ValueError, TypeError):
        return str(value)


def format_date(value: Any, fmt: str = "%Y-%m-%d") -> str:
    """Format date value"""
    if value is None:
        return "-"
    try:
        if isinstance(value, str):
            return value
        if isinstance(value, (date, datetime)):
            return value.strftime(fmt)
        return str(value)
    except Exception:
        return str(value)


def format_days(days: Any) -> str:
    """Format days with appropriate suffix"""
    if days is None:
        return "-"
    try:
        d = int(days)
        if d == 0:
            return "Today"
        elif d == 1:
            return "1 day"
        else:
            return f"{d} days"
    except (ValueError, TypeError):
        return str(days)


# ==================== UI Helpers ====================

def render_category_badge(category: str) -> str:
    """Render HTML badge for category"""
    display = InventoryQualityConstants.CATEGORY_DISPLAY.get(category, category)
    color = InventoryQualityConstants.CATEGORY_COLORS.get(category, '#6c757d')
    return f'<span style="background-color:{color};color:white;padding:2px 8px;border-radius:4px;font-size:0.85em;">{display}</span>'


def render_metric_card(label: str, value: Any, icon: str = "", delta: str = None, color: str = None):
    """Render a metric card with optional delta"""
    st.metric(
        label=f"{icon} {label}" if icon else label,
        value=value,
        delta=delta
    )


def create_excel_download(df: pd.DataFrame, filename: str = "export.xlsx") -> bytes:
    """Create Excel file from DataFrame"""
    from openpyxl.utils import get_column_letter
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Data']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            ) + 2
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length, 50)
    
    return output.getvalue()


# ==================== Session State Helpers ====================

def init_session_state():
    """Initialize session state for Inventory Quality module"""
    defaults = {
        'iq_selected_row': None,
        'iq_show_detail_dialog': False,
        'iq_category_filter': 'All',
        'iq_warehouse_filter': None,
        'iq_product_search': '',
        'iq_data_loaded': False
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def clear_selection():
    """Clear current selection"""
    st.session_state['iq_selected_row'] = None
    st.session_state['iq_show_detail_dialog'] = False


# ==================== Data Validation ====================

def safe_get(data: dict, key: str, default: Any = None) -> Any:
    """Safely get value from dictionary"""
    try:
        value = data.get(key, default)
        if pd.isna(value):
            return default
        return value
    except Exception:
        return default


# ==================== Period Report Helpers ====================

def get_period_dates(preset_key: str) -> tuple:
    """
    Calculate (from_date, to_date) for a given period preset.
    
    Returns:
        Tuple of (from_date, to_date) as date objects
    """
    today = get_vietnam_today()
    
    if preset_key == 'this_month':
        from_date = today.replace(day=1)
        to_date = today
    elif preset_key == 'last_month':
        first_of_current = today.replace(day=1)
        last_month_end = first_of_current - timedelta(days=1)
        from_date = last_month_end.replace(day=1)
        to_date = last_month_end
    elif preset_key == 'this_quarter':
        q = (today.month - 1) // 3
        from_date = date(today.year, q * 3 + 1, 1)
        to_date = today
    elif preset_key == 'last_quarter':
        q = (today.month - 1) // 3
        if q == 0:
            from_date = date(today.year - 1, 10, 1)
            to_date = date(today.year - 1, 12, 31)
        else:
            from_date = date(today.year, (q - 1) * 3 + 1, 1)
            end_month = q * 3
            _, last_day = monthrange(today.year, end_month)
            to_date = date(today.year, end_month, last_day)
    elif preset_key == 'this_year':
        from_date = date(today.year, 1, 1)
        to_date = today
    elif preset_key == 'last_year':
        from_date = date(today.year - 1, 1, 1)
        to_date = date(today.year - 1, 12, 31)
    else:  # custom - default to this month
        from_date = today.replace(day=1)
        to_date = today
    
    return from_date, to_date


def format_report_qty(value: Any, decimals: int = 5) -> str:
    """Format quantity for period report - blank for zero/near-zero"""
    if value is None:
        return ''
    try:
        num = float(value)
        if abs(num) < 0.001:
            return ''
        if num == int(num):
            return f"{int(num):,}"
        return f"{num:,.{decimals}f}".rstrip('0').rstrip('.')
    except (ValueError, TypeError):
        return str(value)


def create_period_summary_excel(df: pd.DataFrame, from_date: date, to_date: date) -> bytes:
    """Create formatted Excel file for period inventory summary"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write data starting from row 4 (rows 1-2 for title, row 3 blank, row 4 header)
        df.to_excel(writer, index=False, sheet_name='Inventory Summary', startrow=3)
        
        ws = writer.sheets['Inventory Summary']
        num_cols = len(df.columns)
        
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1, value='INVENTORY PERIOD SUMMARY')
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Subtitle row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        subtitle_text = f'From {from_date.strftime("%d/%m/%Y")} to {to_date.strftime("%d/%m/%Y")}'
        subtitle_cell = ws.cell(row=2, column=1, value=subtitle_text)
        subtitle_cell.font = Font(italic=True, size=11)
        subtitle_cell.alignment = Alignment(horizontal='center')
        
        # Header styling (row 4)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_border = Border(
            bottom=Side(style='thin', color='2F5496'),
        )
        
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = header_border
        
        # Number formatting for data rows
        num_format = '#,##0.#####'
        for row_idx in range(5, 5 + len(df)):
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Apply number format to quantity columns (typically columns 4-7)
                if col_idx >= 4:
                    cell.number_format = num_format
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(df.columns[col_idx - 1])) + 2
            
            for row in ws.iter_rows(min_row=5, max_row=4 + len(df), 
                                      min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        cell_len = len(str(cell.value or ''))
                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception:
                        pass
            
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    return output.getvalue()