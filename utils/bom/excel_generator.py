# utils/bom/excel_generator.py
"""
Professional Excel Generator for BOM
Creates styled Excel workbook with single comprehensive sheet (like PDF)

VERSION: 3.2.0

CHANGES in v3.2.0:
- Changed legacy code display from "N/A" to "NEW" for products without legacy code
- Format: code (legacy|NEW) | name | pkg (brand)

CHANGES in v3.1.0:
- Updated product display format: code (legacy | N/A) | name | pkg (brand)
- Added legacy_code, package_size, brand to materials and alternatives display

CHANGES in v3.0.0:
- Single sheet layout matching PDF structure
- Company header with name (by language), address, MST
- All information in one sheet: Header, BOM Info, Materials, Alternatives
- Added exported_by parameter
- Vietnamese diacritics removal for English export
"""

import logging
from datetime import datetime
from typing import Dict, Any, Optional, List
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment,
    NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ==================== Vietnamese Diacritics Removal ====================

def remove_vietnamese_diacritics(text: str) -> str:
    """Remove Vietnamese diacritics from text"""
    if not text:
        return text
    
    vietnamese_map = {
        'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'đ': 'd',
        'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        'À': 'A', 'Á': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
        'Ă': 'A', 'Ằ': 'A', 'Ắ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
        'Â': 'A', 'Ầ': 'A', 'Ấ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
        'Đ': 'D',
        'È': 'E', 'É': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
        'Ê': 'E', 'Ề': 'E', 'Ế': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
        'Ì': 'I', 'Í': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
        'Ò': 'O', 'Ó': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
        'Ô': 'O', 'Ồ': 'O', 'Ố': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
        'Ơ': 'O', 'Ờ': 'O', 'Ớ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
        'Ù': 'U', 'Ú': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
        'Ư': 'U', 'Ừ': 'U', 'Ứ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
        'Ỳ': 'Y', 'Ý': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y',
    }
    
    return ''.join(vietnamese_map.get(c, c) for c in text)


# ==================== Product Display Formatting ====================

def format_product_code_with_legacy(code: str, legacy_code: Optional[str] = None) -> str:
    """
    Format product code with legacy code: code (legacy | NEW)
    
    Args:
        code: Product code (pt_code)
        legacy_code: Legacy product code
        
    Returns:
        Formatted string like "VTI001 (OLD-001)" or "VTI001 (NEW)"
    """
    legacy_display = "NEW"
    if legacy_code and str(legacy_code).strip() and str(legacy_code).strip() != '-':
        legacy_display = str(legacy_code).strip()
    
    return f"{code} ({legacy_display})"


def format_product_name_with_details(name: str, 
                                     package_size: Optional[str] = None,
                                     brand: Optional[str] = None) -> str:
    """
    Format product name with package and brand: name | pkg (brand)
    
    Args:
        name: Product name
        package_size: Package size
        brand: Brand name
        
    Returns:
        Formatted string like "Product ABC | 100g (Brand)" or "Product ABC"
    """
    result = name or ""
    
    extra_parts = []
    if package_size and str(package_size).strip() and str(package_size).strip() != '-':
        extra_parts.append(str(package_size).strip())
    
    if brand and str(brand).strip() and str(brand).strip() != '-':
        if extra_parts:
            extra_parts[0] = f"{extra_parts[0]} ({str(brand).strip()})"
        else:
            extra_parts.append(f"({str(brand).strip()})")
    
    if extra_parts:
        result += " | " + " ".join(extra_parts)
    
    return result


# ==================== Style Definitions ====================

# Colors
COMPANY_BG = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
TITLE_BG = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
HEADER_BG = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SUBHEADER_BG = PatternFill(start_color="85929E", end_color="85929E", fill_type="solid")
ALT_ROW_BG = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
INFO_BG = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
SECTION_BG = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")

# Fonts
COMPANY_FONT = Font(name='Arial', size=12, bold=True, color="1E8449")
COMPANY_FONT_SUB = Font(name='Arial', size=10, italic=True, color="1E8449")
TITLE_FONT = Font(name='Arial', size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name='Arial', size=10, bold=True, color="FFFFFF")
SECTION_FONT = Font(name='Arial', size=11, bold=True, color="2C3E50")
LABEL_FONT = Font(name='Arial', size=10, bold=True, color="2C3E50")
VALUE_FONT = Font(name='Arial', size=10, color="000000")
NORMAL_FONT = Font(name='Arial', size=9, color="000000")
FOOTER_FONT = Font(name='Arial', size=8, italic=True, color="7F8C8D")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)

# Alignments
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)


class BOMExcelGenerator:
    """Generate professional Excel workbook for BOM export - Single sheet layout"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        self.company_info = None
        self.language = 'vi'
        self.current_row = 1
    
    def generate(self, bom_info: Dict[str, Any],
                 materials: pd.DataFrame,
                 alternatives_data: Dict[int, pd.DataFrame],
                 company_id: Optional[int] = None,
                 company_info: Optional[Dict] = None,
                 language: str = 'vi',
                 exported_by: Optional[str] = None) -> bytes:
        """
        Generate Excel workbook for BOM - Single comprehensive sheet
        
        Args:
            bom_info: BOM header information
            materials: DataFrame of BOM materials
            alternatives_data: Dict mapping detail_id to alternatives DataFrame
            company_id: Selected company ID (for reference)
            company_info: Pre-fetched company info from export dialog
            language: 'vi' or 'en'
            exported_by: Name of user exporting
            
        Returns:
            Excel file as bytes
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "BOM"
        self.company_info = company_info or {}
        self.language = language
        self.current_row = 1
        
        # Set column widths
        self._set_column_widths()
        
        # Build single sheet with all sections
        self._create_company_header()
        self._create_document_title(bom_info)
        self._create_bom_info_section(bom_info)
        self._create_materials_section(materials)
        self._create_alternatives_section(materials, alternatives_data)
        self._create_footer(bom_info, exported_by)
        
        # Save to bytes
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _set_column_widths(self):
        """Set column widths for optimal display"""
        widths = {
            'A': 6,   # STT
            'B': 18,  # Code
            'C': 45,  # Name
            'D': 15,  # Type
            'E': 12,  # Quantity
            'F': 10,  # UOM
            'G': 10,  # Scrap
            'H': 8,   # Alt
        }
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width
    
    def _create_company_header(self):
        """Create company header section - matching PDF style"""
        if not self.company_info:
            return
        
        # Company name based on language
        if self.language == 'vi':
            company_name = self.company_info.get('local_name') or self.company_info.get('english_name', '')
        else:
            company_name = self.company_info.get('english_name') or self.company_info.get('local_name', '')
        
        # Address (remove diacritics for English)
        address = self.company_info.get('address', '')
        if self.language == 'en' and address:
            address = remove_vietnamese_diacritics(address)
        
        # MST
        mst = self.company_info.get('registration_code', '')
        
        # Row 1: Company name
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=company_name)
        cell.font = COMPANY_FONT
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.fill = COMPANY_BG
        self.ws.row_dimensions[self.current_row].height = 22
        self.current_row += 1
        
        # Row 2: Address
        if address:
            self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
            cell = self.ws.cell(row=self.current_row, column=1, value=address)
            cell.font = COMPANY_FONT_SUB
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = COMPANY_BG
            self.current_row += 1
        
        # Row 3: MST
        if mst:
            self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
            cell = self.ws.cell(row=self.current_row, column=1, value=f"MST: {mst}")
            cell.font = COMPANY_FONT_SUB
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = COMPANY_BG
            self.current_row += 1
        
        # Empty row
        self.current_row += 1
    
    def _create_document_title(self, bom_info: Dict):
        """Create document title section"""
        # Title row
        if self.language == 'vi':
            title = "ĐỊNH MỨC SẢN XUẤT (BOM)"
        else:
            title = "BILL OF MATERIALS (BOM)"
        
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=title)
        cell.font = TITLE_FONT
        cell.fill = TITLE_BG
        cell.alignment = CENTER_ALIGN
        self.ws.row_dimensions[self.current_row].height = 28
        self.current_row += 2
        
        # BOM Code subtitle
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=bom_info.get('bom_code', ''))
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.alignment = CENTER_ALIGN
        self.current_row += 2
    
    def _create_bom_info_section(self, bom_info: Dict):
        """Create BOM information section"""
        # Section header
        if self.language == 'vi':
            section_title = "📋 THÔNG TIN BOM"
        else:
            section_title = "📋 BOM INFORMATION"
        
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=section_title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_BG
        self.current_row += 1
        
        # Labels based on language
        if self.language == 'vi':
            labels = {
                'code': 'Mã BOM', 'name': 'Tên BOM', 'type': 'Loại BOM', 'status': 'Trạng thái',
                'product': 'Sản phẩm', 'output': 'Sản lượng', 'effective': 'Ngày hiệu lực',
                'version': 'Phiên bản', 'creator': 'Người tạo', 'created_date': 'Ngày tạo'
            }
        else:
            labels = {
                'code': 'BOM Code', 'name': 'BOM Name', 'type': 'BOM Type', 'status': 'Status',
                'product': 'Output Product', 'output': 'Output Qty', 'effective': 'Effective Date',
                'version': 'Version', 'creator': 'Created By', 'created_date': 'Created Date'
            }
        
        # Format values
        product_display = f"{bom_info.get('product_code', '')} - {bom_info.get('product_name', '')}"
        output_qty = bom_info.get('output_qty', 0)
        output_str = f"{output_qty:,.2f} {bom_info.get('uom', 'PCS')}"
        
        effective_date = bom_info.get('effective_date', '')
        if effective_date and hasattr(effective_date, 'strftime'):
            effective_date = effective_date.strftime('%d/%m/%Y')
        
        created_date = bom_info.get('created_date', '')
        if created_date and hasattr(created_date, 'strftime'):
            created_date = created_date.strftime('%d/%m/%Y %H:%M')
        
        creator_name = bom_info.get('creator_name', 'Unknown')
        
        # Info rows - 2 columns layout
        info_rows = [
            (labels['code'], bom_info.get('bom_code', ''), labels['product'], product_display),
            (labels['name'], bom_info.get('bom_name', ''), labels['output'], output_str),
            (labels['type'], bom_info.get('bom_type', ''), labels['effective'], str(effective_date) or 'N/A'),
            (labels['status'], bom_info.get('status', ''), labels['version'], str(bom_info.get('version', 1))),
            (labels['creator'], creator_name, labels['created_date'], str(created_date) or 'N/A'),
        ]
        
        for label1, value1, label2, value2 in info_rows:
            # Left column - Label (A)
            cell = self.ws.cell(row=self.current_row, column=1, value=label1)
            cell.font = LABEL_FONT
            cell.fill = INFO_BG
            cell.border = THIN_BORDER
            
            # Left column - Value (B-C)
            self.ws.merge_cells(f'B{self.current_row}:C{self.current_row}')
            cell = self.ws.cell(row=self.current_row, column=2, value=value1)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            
            # Right column - Label (D)
            cell = self.ws.cell(row=self.current_row, column=4, value=label2)
            cell.font = LABEL_FONT
            cell.fill = INFO_BG
            cell.border = THIN_BORDER
            
            # Right column - Value (E-H)
            self.ws.merge_cells(f'E{self.current_row}:H{self.current_row}')
            cell = self.ws.cell(row=self.current_row, column=5, value=value2)
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            
            self.current_row += 1
        
        # Notes if exist
        if bom_info.get('notes'):
            note_label = 'Ghi chú' if self.language == 'vi' else 'Notes'
            cell = self.ws.cell(row=self.current_row, column=1, value=note_label)
            cell.font = LABEL_FONT
            cell.fill = INFO_BG
            cell.border = THIN_BORDER
            
            self.ws.merge_cells(f'B{self.current_row}:H{self.current_row}')
            cell = self.ws.cell(row=self.current_row, column=2, value=bom_info.get('notes', ''))
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            self.current_row += 1
        
        self.current_row += 1
    
    def _create_materials_section(self, materials: pd.DataFrame):
        """Create materials list section"""
        # Section header
        if self.language == 'vi':
            section_title = "🧱 DANH SÁCH NGUYÊN VẬT LIỆU"
        else:
            section_title = "🧱 MATERIALS LIST"
        
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=section_title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_BG
        self.current_row += 1
        
        if materials.empty:
            self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
            msg = "Không có nguyên vật liệu" if self.language == 'vi' else "No materials"
            self.ws.cell(row=self.current_row, column=1, value=msg)
            self.current_row += 2
            return
        
        # Table headers
        if self.language == 'vi':
            headers = ['STT', 'Mã NVL', 'Tên nguyên vật liệu', 'Loại', 'Số lượng', 'ĐVT', 'Hao hụt', 'Alt']
        else:
            headers = ['#', 'Code', 'Material Name', 'Type', 'Quantity', 'UOM', 'Scrap', 'Alt']
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_BG
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        
        self.ws.row_dimensions[self.current_row].height = 22
        self.current_row += 1
        
        # Data rows
        for idx, (_, mat) in enumerate(materials.iterrows(), 1):
            fill = ALT_ROW_BG if idx % 2 == 0 else None
            
            # Format quantity
            qty = float(mat.get('quantity', 0))
            if qty == int(qty):
                qty_str = f"{int(qty):,}"
            else:
                qty_str = f"{qty:,.4f}".rstrip('0').rstrip('.')
            
            # Format scrap
            scrap = float(mat.get('scrap_rate', 0))
            scrap_str = f"{scrap:.1f}%" if scrap > 0 else "-"
            
            # Alt count
            alt_count = int(mat.get('alternatives_count', 0))
            alt_str = str(alt_count) if alt_count > 0 else "-"
            
            # Format code with legacy: code (legacy | N/A)
            code_display = format_product_code_with_legacy(
                mat.get('material_code', ''),
                mat.get('legacy_code')
            )
            
            # Format name with details: name | pkg (brand)
            name_display = format_product_name_with_details(
                mat.get('material_name', ''),
                mat.get('package_size'),
                mat.get('brand')
            )
            
            row_data = [
                idx,
                code_display,
                name_display,
                mat.get('material_type', ''),
                qty_str,
                mat.get('uom', ''),
                scrap_str,
                alt_str
            ]
            
            alignments = [CENTER_ALIGN, LEFT_ALIGN, LEFT_ALIGN, CENTER_ALIGN,
                         RIGHT_ALIGN, CENTER_ALIGN, CENTER_ALIGN, CENTER_ALIGN]
            
            for col, (value, align) in enumerate(zip(row_data, alignments), 1):
                cell = self.ws.cell(row=self.current_row, column=col, value=value)
                cell.font = NORMAL_FONT
                cell.alignment = align
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill
            
            self.current_row += 1
        
        self.current_row += 1
    
    def _create_alternatives_section(self, materials: pd.DataFrame,
                                      alternatives_data: Dict[int, pd.DataFrame]):
        """Create alternatives section if any exist"""
        # Check if any alternatives exist
        has_alternatives = False
        for _, mat in materials.iterrows():
            detail_id = int(mat['id'])
            if detail_id in alternatives_data and not alternatives_data[detail_id].empty:
                has_alternatives = True
                break
        
        if not has_alternatives:
            return
        
        # Section header
        if self.language == 'vi':
            section_title = "🔄 NGUYÊN VẬT LIỆU THAY THẾ"
        else:
            section_title = "🔄 ALTERNATIVE MATERIALS"
        
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=section_title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_BG
        self.current_row += 1
        
        # Process each material with alternatives
        for _, mat in materials.iterrows():
            detail_id = int(mat['id'])
            alternatives = alternatives_data.get(detail_id)
            
            if alternatives is None or alternatives.empty:
                continue
            
            # Material header
            mat_title = f"▸ {mat['material_code']} - {mat['material_name']}"
            self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
            cell = self.ws.cell(row=self.current_row, column=1, value=mat_title)
            cell.font = Font(name='Arial', size=9, bold=True, color="2C3E50")
            self.current_row += 1
            
            # Alternatives table headers
            if self.language == 'vi':
                alt_headers = ['ƯT', 'Mã NVL thay thế', 'Tên nguyên vật liệu thay thế', '', 'SL', 'ĐVT', 'Hao hụt', 'TT']
            else:
                alt_headers = ['P', 'Alt Code', 'Alternative Material Name', '', 'Qty', 'UOM', 'Scrap', 'Status']
            
            for col, header in enumerate(alt_headers, 1):
                cell = self.ws.cell(row=self.current_row, column=col, value=header)
                cell.font = Font(name='Arial', size=8, bold=True, color="FFFFFF")
                cell.fill = SUBHEADER_BG
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            
            # Merge columns C and D for name
            self.ws.merge_cells(f'C{self.current_row}:D{self.current_row}')
            self.current_row += 1
            
            # Alternatives data
            for _, alt in alternatives.iterrows():
                status_icon = "✓" if alt['is_active'] else "○"
                
                qty = float(alt['quantity'])
                if qty == int(qty):
                    qty_str = f"{int(qty):,}"
                else:
                    qty_str = f"{qty:,.2f}".rstrip('0').rstrip('.')
                
                scrap = float(alt.get('scrap_rate', 0))
                scrap_str = f"{scrap:.1f}%" if scrap > 0 else "-"
                
                # Format code with legacy: code (legacy | N/A)
                alt_code_display = format_product_code_with_legacy(
                    alt.get('material_code', ''),
                    alt.get('legacy_code')
                )
                
                # Format name with details: name | pkg (brand)
                alt_name_display = format_product_name_with_details(
                    alt.get('material_name', ''),
                    alt.get('package_size'),
                    alt.get('brand')
                )
                
                alt_row_data = [
                    alt['priority'],
                    alt_code_display,
                    alt_name_display,
                    '',  # Merged with C
                    qty_str,
                    alt['uom'],
                    scrap_str,
                    status_icon
                ]
                
                for col, value in enumerate(alt_row_data, 1):
                    cell = self.ws.cell(row=self.current_row, column=col, value=value)
                    cell.font = Font(name='Arial', size=8)
                    cell.border = THIN_BORDER
                    if col in [1, 5, 6, 7, 8]:
                        cell.alignment = CENTER_ALIGN
                    elif col == 5:
                        cell.alignment = RIGHT_ALIGN
                    else:
                        cell.alignment = LEFT_ALIGN
                
                # Merge columns C and D for name
                self.ws.merge_cells(f'C{self.current_row}:D{self.current_row}')
                self.current_row += 1
            
            self.current_row += 1
    
    def _create_footer(self, bom_info: Dict, exported_by: Optional[str]):
        """Create footer with export information"""
        self.current_row += 1
        
        # Separator line
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value='─' * 80)
        cell.font = Font(name='Arial', size=6, color="BDC3C7")
        cell.alignment = CENTER_ALIGN
        self.current_row += 1
        
        # Footer info
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        version = bom_info.get('version', 1)
        status = bom_info.get('status', 'N/A')
        
        if self.language == 'vi':
            if exported_by:
                footer_text = f"Xuất bởi: {exported_by} | Ngày xuất: {timestamp} | Phiên bản: {version} | Trạng thái: {status}"
            else:
                footer_text = f"Ngày xuất: {timestamp} | Phiên bản: {version} | Trạng thái: {status}"
        else:
            if exported_by:
                footer_text = f"Exported by: {exported_by} | Generated: {timestamp} | Version: {version} | Status: {status}"
            else:
                footer_text = f"Generated: {timestamp} | Version: {version} | Status: {status}"
        
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws.cell(row=self.current_row, column=1, value=footer_text)
        cell.font = FOOTER_FONT
        cell.alignment = CENTER_ALIGN
    
    def _format_datetime(self, dt) -> str:
        """Format datetime for display"""
        if dt is None:
            return 'N/A'
        if hasattr(dt, 'strftime'):
            return dt.strftime('%d/%m/%Y %H:%M:%S')
        return str(dt) if dt else 'N/A'


# ==================== Helper Function ====================

def generate_bom_excel(bom_info: Dict[str, Any],
                       materials: pd.DataFrame,
                       alternatives_data: Dict[int, pd.DataFrame],
                       company_id: Optional[int] = None,
                       company_info: Optional[Dict] = None,
                       language: str = 'vi',
                       exported_by: Optional[str] = None) -> bytes:
    """
    Convenience function to generate BOM Excel
    
    Args:
        bom_info: BOM header information
        materials: DataFrame of BOM materials
        alternatives_data: Dict mapping detail_id to alternatives DataFrame
        company_id: Selected company ID (for reference)
        company_info: Pre-fetched company info from export dialog
        language: 'vi' or 'en'
        exported_by: Name of user exporting
        
    Returns:
        Excel file as bytes
    """
    generator = BOMExcelGenerator()
    return generator.generate(
        bom_info, materials, alternatives_data,
        company_id=company_id,
        company_info=company_info,
        language=language,
        exported_by=exported_by
    )