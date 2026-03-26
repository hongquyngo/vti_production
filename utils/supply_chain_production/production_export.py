# utils/supply_chain_production/production_export.py

"""
Export Module for Production Planning — multi-sheet formatted Excel.

Sheets:
1. Ready MOs — items ready to start production
2. Waiting MOs — partial materials, with bottleneck info
3. Blocked MOs — no materials / no ETA
4. Unschedulable — missing config items
5. Material Matrix — readiness matrix (product × material)
6. Summary — KPIs, reconciliation, config snapshot
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
import logging
from typing import Optional

from .production_constants import URGENCY_LEVELS, READINESS_STATUS, VERSION
from .mo_result import MOSuggestionResult

logger = logging.getLogger(__name__)

HEADER_FILL_HEX = '2E75B6'
HEADER_FONT_COLOR = 'FFFFFF'
URGENCY_COLORS = {
    'OVERDUE': 'F2DCDB', 'CRITICAL': 'F2DCDB',
    'URGENT': 'FDE9D9', 'THIS_WEEK': 'FFFFCC', 'PLANNED': 'D6EAFF',
}
READINESS_COLORS = {
    'READY': 'C6EFCE', 'USE_ALTERNATIVE': 'C6EFCE',
    'PARTIAL_READY': 'FFEB9C', 'BLOCKED': 'FFC7CE',
}
DEFAULT_COL_WIDTHS = {
    'Priority': 10, 'Urgency': 14, 'Code': 14, 'Product': 30, 'Brand': 12,
    'BOM Type': 12, 'Shortage Qty': 14, 'Suggested Qty': 14, 'Batches': 10,
    'UOM': 8, 'Readiness': 16, 'Materials Ready': 14, 'Max Producible Now': 16,
    'Bottleneck Material': 18, 'Bottleneck ETA': 14, 'Contention': 12,
    'Demand Date': 14, 'Must Start By': 14, 'Start Date': 14,
    'Expected Completion': 16, 'Lead Time (days)': 14, 'LT Source': 14,
    'Delayed': 10, 'Delay Days': 12, 'Delay Reason': 18,
    'At Risk Value ($)': 16, 'Customers': 12, 'Has SO': 10,
    'Existing MOs': 14, 'Existing Remaining': 16,
    'Action': 18, 'Action Detail': 40,
    'Yield Multiplier': 14, 'Yield Source': 12,
    # Material matrix
    'Material Code': 14, 'Material Name': 25, 'Required Qty': 14,
    'Available Now': 14, 'Allocated Qty': 14, 'Coverage %': 12,
    'Status': 12, 'Primary': 10, 'Contested': 10,
    'ETA': 14, 'Coverage Source': 14,
    # Summary
    'Metric': 35, 'Value': 40,
    # Unschedulable
    'Reason': 20, 'Detail': 40, 'Fix': 40,
}


def _lines_to_df(lines, include_readiness=True):
    """Convert MOLineItems to export DataFrame."""
    rows = []
    for l in lines:
        urgency_cfg = URGENCY_LEVELS.get(l.urgency_level, {})
        readiness_cfg = READINESS_STATUS.get(l.readiness_status, {})

        row = {
            'Priority': round(l.priority_score, 1),
            'Urgency': urgency_cfg.get('label', l.urgency_level),
            'Code': l.pt_code,
            'Product': l.product_name,
            'Brand': l.brand,
            'BOM Type': l.bom_type,
            'Shortage Qty': round(l.shortage_qty),
            'Suggested Qty': round(l.suggested_qty),
            'Batches': l.batches_needed,
            'UOM': l.uom,
        }

        if include_readiness:
            row['Readiness'] = readiness_cfg.get('label', l.readiness_status)
            row['Materials Ready'] = f"{l.ready_materials}/{l.total_materials}"
            row['Max Producible Now'] = round(l.max_producible_now) if l.max_producible_now > 0 else ''
            row['Bottleneck Material'] = l.bottleneck_material or ''
            row['Bottleneck ETA'] = str(l.bottleneck_eta) if l.bottleneck_eta else ''
            row['Contention'] = 'Yes' if l.has_contention else ''

        row['Demand Date'] = str(l.demand_date) if l.demand_date else ''
        row['Must Start By'] = str(l.must_start_by) if l.must_start_by else ''
        row['Start Date'] = str(l.actual_start) if l.actual_start else ''
        row['Expected Completion'] = str(l.expected_completion) if l.expected_completion else ''
        row['Lead Time (days)'] = l.lead_time_days
        row['LT Source'] = l.lead_time_source
        row['Delayed'] = 'Yes' if l.is_delayed else ''
        if l.is_delayed and l.delay_reason == 'MATERIAL_BLOCKED_NO_ETA':
            row['Delay Days'] = 'N/A'
        else:
            row['Delay Days'] = l.delay_days if l.is_delayed else ''
        row['Delay Reason'] = l.delay_reason if l.is_delayed else ''
        row['At Risk Value ($)'] = round(l.at_risk_value, 2)
        row['Customers'] = l.customer_count
        row['Has SO'] = 'Yes' if l.has_sales_order else ''
        row['Existing MOs'] = l.existing_mo_count if l.existing_mo_count > 0 else ''
        row['Existing Remaining'] = round(l.existing_mo_remaining_qty) if l.existing_mo_remaining_qty > 0 else ''
        row['Yield Multiplier'] = round(l.yield_multiplier, 4) if l.yield_multiplier != 1.0 else ''
        row['Yield Source'] = l.yield_source if l.yield_source else ''
        row['Action'] = l.action_type
        row['Action Detail'] = l.action_description

        rows.append(row)

    return pd.DataFrame(rows)


def _apply_sheet_formatting(ws, df, col_widths=None, currency_cols=None,
                            number_cols=None, date_cols=None,
                            urgency_col=None, readiness_col=None):
    """Apply professional formatting to a worksheet."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    hdr_font = Font(name='Arial', bold=True, color=HEADER_FONT_COLOR, size=10)
    hdr_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dfont = Font(name='Arial', size=10)

    nrows, ncols = len(df), len(df.columns)

    # Header row
    for ci in range(1, ncols + 1):
        c = ws.cell(row=1, column=ci)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, thin

    ws.freeze_panes = 'A2'
    if nrows > 0:
        ws.auto_filter.ref = ws.dimensions

    col_map = {name: idx + 1 for idx, name in enumerate(df.columns)}
    widths = dict(DEFAULT_COL_WIDTHS)
    if col_widths:
        widths.update(col_widths)

    for col_name, ci in col_map.items():
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = widths.get(col_name, 14)

    # Data formatting
    for ri in range(2, nrows + 2):
        for ci in range(1, ncols + 1):
            c = ws.cell(row=ri, column=ci)
            c.font = dfont
            c.border = thin

    # Currency formatting
    if currency_cols:
        for col_name in currency_cols:
            if col_name in col_map:
                ci = col_map[col_name]
                for ri in range(2, nrows + 2):
                    ws.cell(row=ri, column=ci).number_format = '$#,##0.00'

    # Number formatting
    if number_cols:
        for col_name in number_cols:
            if col_name in col_map:
                ci = col_map[col_name]
                for ri in range(2, nrows + 2):
                    ws.cell(row=ri, column=ci).number_format = '#,##0'

    # Urgency row coloring
    if urgency_col and urgency_col in col_map:
        ci = col_map[urgency_col]
        for ri in range(2, nrows + 2):
            val = ws.cell(row=ri, column=ci).value
            color = URGENCY_COLORS.get(val)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                for cj in range(1, ncols + 1):
                    ws.cell(row=ri, column=cj).fill = fill

    # Readiness row coloring
    if readiness_col and readiness_col in col_map:
        ci = col_map[readiness_col]
        for ri in range(2, nrows + 2):
            val = ws.cell(row=ri, column=ci).value
            # Map label back to key
            label_to_key = {v.get('label', k): k for k, v in READINESS_STATUS.items()}
            key = label_to_key.get(val, val)
            color = READINESS_COLORS.get(key)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                for cj in range(1, ncols + 1):
                    ws.cell(row=ri, column=cj).fill = fill


def export_mo_suggestions_to_excel(result: MOSuggestionResult) -> BytesIO:
    """Export full MO suggestion result to multi-sheet Excel."""
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Sheet 1: Ready MOs
        ready_df = _lines_to_df(result.ready_lines, include_readiness=False)
        if ready_df.empty:
            ready_df = pd.DataFrame({'Info': ['No items ready to produce']})
        ready_df.to_excel(writer, sheet_name='Ready MOs', index=False)
        _apply_sheet_formatting(
            writer.sheets['Ready MOs'], ready_df,
            currency_cols=['At Risk Value ($)'],
            number_cols=['Shortage Qty', 'Suggested Qty', 'Batches'],
            urgency_col='Urgency',
        )

        # Sheet 2: Waiting MOs
        waiting_df = _lines_to_df(result.waiting_lines, include_readiness=True)
        if waiting_df.empty:
            waiting_df = pd.DataFrame({'Info': ['No items waiting for materials']})
        waiting_df.to_excel(writer, sheet_name='Waiting MOs', index=False)
        _apply_sheet_formatting(
            writer.sheets['Waiting MOs'], waiting_df,
            currency_cols=['At Risk Value ($)'],
            number_cols=['Shortage Qty', 'Suggested Qty'],
            readiness_col='Readiness',
        )

        # Sheet 3: Blocked MOs
        blocked_df = _lines_to_df(result.blocked_lines, include_readiness=True)
        if blocked_df.empty:
            blocked_df = pd.DataFrame({'Info': ['No blocked items']})
        blocked_df.to_excel(writer, sheet_name='Blocked MOs', index=False)
        _apply_sheet_formatting(
            writer.sheets['Blocked MOs'], blocked_df,
            currency_cols=['At Risk Value ($)'],
            number_cols=['Shortage Qty', 'Suggested Qty'],
        )

        # Sheet 4: Unschedulable
        unsch_df = result.get_unschedulable_df()
        if unsch_df.empty:
            unsch_df = pd.DataFrame({'Info': ['All items schedulable']})
        else:
            unsch_df.columns = [c.replace('_', ' ').title() for c in unsch_df.columns]
        unsch_df.to_excel(writer, sheet_name='Unschedulable', index=False)
        _apply_sheet_formatting(
            writer.sheets['Unschedulable'], unsch_df,
            number_cols=['Shortage Qty'],
        )

        # Sheet 5: Material Matrix
        mat_df = result.get_readiness_matrix_df()
        if mat_df.empty:
            mat_df = pd.DataFrame({'Info': ['No material readiness data']})
        else:
            mat_df.columns = [c.replace('_', ' ').title() for c in mat_df.columns]
        mat_df.to_excel(writer, sheet_name='Material Matrix', index=False)
        _apply_sheet_formatting(
            writer.sheets['Material Matrix'], mat_df,
            number_cols=['Required Qty', 'Available Now', 'Allocated Qty'],
        )

        # Sheet 6: Summary
        summary_rows = _build_summary_rows(result)
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        _apply_sheet_formatting(writer.sheets['Summary'], summary_df)

    buffer.seek(0)
    return buffer


def _build_summary_rows(result: MOSuggestionResult):
    """Build summary metrics as key-value rows."""
    m = result.get_summary()
    recon = result.get_reconciliation()
    config = result.config_snapshot or {}

    rows = [
        {'Metric': 'Report Generated', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M')},
        {'Metric': 'Module Version', 'Value': VERSION},
        {'Metric': '', 'Value': ''},
        {'Metric': '── Results ──', 'Value': ''},
        {'Metric': 'Total MO Suggestions', 'Value': m.get('total_mo_lines', 0)},
        {'Metric': 'Ready to Produce', 'Value': m.get('ready_count', 0)},
        {'Metric': 'Waiting for Materials', 'Value': m.get('waiting_count', 0)},
        {'Metric': 'Blocked', 'Value': m.get('blocked_count', 0)},
        {'Metric': 'Unschedulable', 'Value': m.get('unschedulable_count', 0)},
        {'Metric': '', 'Value': ''},
        {'Metric': '── Values ──', 'Value': ''},
        {'Metric': 'Total At-Risk Value', 'Value': f"${m.get('total_at_risk_value', 0):,.0f}"},
        {'Metric': 'Ready At-Risk Value', 'Value': f"${m.get('ready_at_risk_value', 0):,.0f}"},
        {'Metric': 'Waiting At-Risk Value', 'Value': f"${m.get('waiting_at_risk_value', 0):,.0f}"},
        {'Metric': 'Blocked At-Risk Value', 'Value': f"${m.get('blocked_at_risk_value', 0):,.0f}"},
        {'Metric': '', 'Value': ''},
        {'Metric': '── Urgency ──', 'Value': ''},
        {'Metric': 'Overdue', 'Value': m.get('overdue_count', 0)},
        {'Metric': 'Delayed', 'Value': m.get('delayed_count', 0)},
        {'Metric': 'Contention', 'Value': m.get('contention_count', 0)},
        {'Metric': '', 'Value': ''},
        {'Metric': '── Reconciliation ──', 'Value': ''},
        {'Metric': 'Total Input', 'Value': recon.get('total_input', 0)},
        {'Metric': 'Skipped (Validation)', 'Value': recon.get('input_skipped_validation', 0)},
        {'Metric': 'Total Accounted', 'Value': recon.get('total_accounted', 0)},
        {'Metric': 'Balanced', 'Value': '✅ Yes' if recon.get('is_balanced') else '❌ No'},
        {'Metric': '', 'Value': ''},
        {'Metric': '── Config Snapshot ──', 'Value': ''},
    ]

    for key, val in config.items():
        rows.append({'Metric': key, 'Value': str(val)})

    return rows


def get_mo_export_filename() -> str:
    """Generate timestamped filename for MO export."""
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    return f"MO_Suggestions_{ts}.xlsx"
